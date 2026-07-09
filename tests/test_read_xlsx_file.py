from pathlib import Path

import openpyxl
import pytest

from piidigger.filehandlers.xlsx import XlsxHandler
from piidigger.models.config import Config, SpreadsheetConfig
from piidigger.orchestration.sources import FilesystemItem


def _read(path: Path, config: Config | None = None) -> list[str]:
    return list(XlsxHandler().read(FilesystemItem(path), config or Config()))


@pytest.mark.filehandlers
def test_xlsx_missing_file() -> None:
    with pytest.raises(FileNotFoundError):
        _read(Path("testdata/xlsx/does-not-exist.xlsx"))


@pytest.mark.filehandlers
def test_xlsx_empty_file() -> None:
    chunks = _read(Path("testdata/xlsx/empty-file.xlsx"))
    assert chunks == []


# Small, predictable fixtures: exact per-sheet chunk assertions.
# XlsxHandler yields one chunk per sheet; with the default buffer size each sheet's
# entire content fits in a single chunk.
@pytest.mark.filehandlers
@pytest.mark.parametrize(
    "filename, expected_chunks",
    [
        (
            "testdata/xlsx/test-1sheet-1cell.xlsx",
            ["S1R1C1"],
        ),
        (
            "testdata/xlsx/test-1sheet-1cell-carriage-return.xlsx",
            ["S1R1C1L1 S1R1C1L2"],
        ),
        (
            "testdata/xlsx/test-1sheet-1row.xlsx",
            ["S1R1C1 S1R1C2 S1R1C3 S1R1C4"],
        ),
        (
            "testdata/xlsx/test-1sheet-10row-table.xlsx",
            [
                "Sheet Row Column Text 1 1 3 S1R1C3 1 2 3 S1R2C3 1 3 3 S1R3C3 1 4 3 S1R4C3 1 5 3 S1R5C3 1 6 3 S1R6C3 1 7 3 S1R7C3 1 8 3 S1R8C3 1 9 3 S1R9C3 1 10 3 S1R10C3"
            ],
        ),
        (
            "testdata/xlsx/test-2sheet-10row-table.xlsx",
            [
                "Sheet Row Column Text 1 1 3 S1R1C3 1 2 3 S1R2C3 1 3 3 S1R3C3 1 4 3 S1R4C3 1 5 3 S1R5C3 1 6 3 S1R6C3 1 7 3 S1R7C3 1 8 3 S1R8C3 1 9 3 S1R9C3 1 10 3 S1R10C3",
                "Sheet Row Column Text 1 1 3 S1R1C3 1 2 3 S1R2C3 1 3 3 S1R3C3 1 4 3 S1R4C3 1 5 3 S1R5C3 1 6 3 S1R6C3 1 7 3 S1R7C3 1 8 3 S1R8C3 1 9 3 S1R9C3 1 10 3 S1R10C3",
            ],
        ),
    ],
)
def test_xlsx_exact_content(filename: str, expected_chunks: list[str]) -> None:
    chunks = _read(Path(filename))
    assert chunks == expected_chunks


@pytest.mark.filehandlers
def test_xlsx_random_data_table() -> None:
    # Large table that was split into 22 chunks with maxChunkCount=2; now
    # arrives as a single chunk with the default buffer size.
    chunks = _read(Path("testdata/xlsx/random-data-table.xlsx"))
    content = " ".join(chunks)
    assert "First Name" in content
    assert "j.montgomery@randatmail.com" in content
    assert "Lower secondary" in content


@pytest.mark.filehandlers
def test_xlsx_blank_row_limit_stops_early(tmp_path: Path) -> None:
    # One value, a run of blank rows, then a second value further down.
    workbook_path = tmp_path / "blank-run.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "FIRST"
    ws["A10"] = "SECOND"  # 8 blank rows (2-9) separate the two values
    wb.save(workbook_path)

    default_content = " ".join(_read(workbook_path))
    assert "FIRST" in default_content
    assert "SECOND" in default_content

    strict_config = Config(spreadsheet=SpreadsheetConfig(blank_row_limit=3))
    strict_content = " ".join(_read(workbook_path, strict_config))
    assert "FIRST" in strict_content
    assert "SECOND" not in strict_content
