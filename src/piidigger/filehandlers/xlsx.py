import warnings
from collections.abc import Iterator
from io import BytesIO

import openpyxl
from openpyxl.cell.cell import MergedCell

from piidigger.filehandlers._sharedfuncs import ContentBuffer
from piidigger.models.config import Config

# Ignore the UserWarning message from OpenPyXL that seem to pop up here and there
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

HANDLES = {
    "ext": [
        ".xlsx",
        ".xlsm",
        ".xlst",
        ".xltm",
    ],
    "mime": [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroEnabled",
        "application/vnd.ms-excel.template",
    ],
}

handles = HANDLES


class XlsxHandler:
    """FileHandler for XLSX/XLSM/XLTM files.

    Preferred path (archive members): source.open_bytes() returns bytes which
    are wrapped in BytesIO and passed to load_workbook with read_only=False.
    Members are bounded by max_member_uncompressed_size_mb so full buffering
    is safe.

    Fallback path (on-disk files): source.open_bytes() returns None, so
    source.materialize() is called for a filesystem path and read_only=True
    is used for memory-efficient streaming of potentially large files.
    """

    def read(self, source, config: Config) -> Iterator[str]:  # source: ScannableItem
        data = source.open_bytes()
        if data is not None:
            book = openpyxl.load_workbook(filename=BytesIO(data), read_only=False, data_only=True)
        else:
            book = openpyxl.load_workbook(filename=str(source.materialize()), read_only=True, data_only=True)
        try:
            for sheet_name in book.sheetnames:
                active_sheet = book[sheet_name]
                content_buffer: ContentBuffer = ContentBuffer(max_bytes=config.buffer.max_buffer_bytes)
                blank_row_count = 0
                row_count = 0

                for row in active_sheet.iter_rows(values_only=True):
                    row_count += 1
                    row_has_data = False
                    line = ""
                    blank_col_count = 0

                    for item in row:
                        if isinstance(item, MergedCell):
                            continue
                        if item is None or item == "":
                            blank_col_count += 1
                            if blank_col_count > config.spreadsheet.blank_col_limit:
                                break
                            continue
                        line += str(item) + " "
                        row_has_data = True

                    content_buffer.append_content(line)
                    if row_has_data:
                        blank_row_count = 0
                    else:
                        blank_row_count += 1
                        if blank_row_count > config.spreadsheet.blank_row_limit:
                            break

                    if content_buffer.content_buffer_full():
                        yield content_buffer.get_content()

                final = content_buffer.finalize_content()
                if final:
                    yield final
        finally:
            book.close()


handler = XlsxHandler()
